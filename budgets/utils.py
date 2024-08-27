import openpyxl
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, PatternFill
from django.shortcuts import get_object_or_404
from django.http import HttpResponse
from .models import Budget, BudgetItem
from projects.models import Contractor, Client, Project
from decimal import Decimal
from django.conf import settings
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill

def export_budget_report_to_excel(request, pk):
    budget = get_object_or_404(Budget, pk=pk)
    project = budget.project
    client = project.client
    contractor = project.contractor
    
    # Cargar la plantilla de Excel
    file_path = f'{settings.BASE_DIR}/static/plantilla.xlsx'
    plantilla = openpyxl.load_workbook(file_path)
    hoja_cotizacion = plantilla['COTIZACION']

    # Llenar los datos del contratista
    hoja_cotizacion['C13'] = contractor.contractor_name
    hoja_cotizacion['C14'] = contractor.contractor_ruc
    hoja_cotizacion['C15'] = contractor.address

    # Llenar los datos del cliente
    hoja_cotizacion['C18'] = client.client_name
    hoja_cotizacion['C20'] = client.email
    hoja_cotizacion['C21'] = client.phone

    # Llenar los datos del proyecto
    hoja_cotizacion['C27'] = budget.budget_name
    hoja_cotizacion['G30'] = budget.total_dolares_final

    # Nuevos campos a llenar según tu solicitud
    hoja_cotizacion['E11'] = budget.numero_cotizacion
    hoja_cotizacion['D33'] = budget.moneda
    hoja_cotizacion['D34'] = f'{budget.impuesto}%'
    hoja_cotizacion['D35'] = budget.tiempo_entrega
    hoja_cotizacion['D36'] = budget.tiempo_servicio
    hoja_cotizacion['D37'] = budget.facturacion
    hoja_cotizacion['D38'] = budget.tiempo_garantia
    hoja_cotizacion['D39'] = "30 días"
    hoja_cotizacion['B40'] = budget.no_incluye
    hoja_cotizacion['B41'] = ""
    hoja_cotizacion['G20'] = budget.numero_cotizacion
    hoja_cotizacion['G21'] = budget.fecha

    # Estilos generales
    header_font = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
    cell_font = Font(name="Calibri", size=11)
    alignment_center = Alignment(horizontal="center", vertical="center")
    fill_blue = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )

    # Agrupar los ítems por tipo (content_type)
    items_by_type = {}
    for item in budget.items.all():
        tipo = item.content_type.model
        if tipo not in items_by_type:
            items_by_type[tipo] = []
        items_by_type[tipo].append(item)

    # Crear una hoja para cada tipo de ítem sin resumen financiero
    for tipo, items in items_by_type.items():
        ws_budget = plantilla.create_sheet(title=tipo.capitalize())

        # Título del presupuesto
        ws_budget['A1'] = f'PRESUPUESTO: {budget.budget_name} ({tipo.capitalize()})'
        ws_budget['A1'].font = Font(bold=True, size=16, name="Calibri")
        ws_budget['A1'].alignment = alignment_center

        # Encabezados de la tabla
        headers = ['ITEM', 'DESCRIPCION DE ITEM', 'UND', 'CANT', 'P.U.', 'SUBTOTAL']
        ws_budget.append(headers)
        for cell in ws_budget[2]:  # Segunda fila
            cell.font = header_font
            cell.fill = fill_blue
            cell.alignment = alignment_center
            cell.border = thin_border

        # Datos de los ítems
        for item in items:
            precio_total_dias = Decimal(item.precio_item_proyecto)
            row = [
                item.inventario.num_articulo,
                item.inventario.descripcion,
                item.unidad_medida,
                item.cantidad,
                Decimal(item.real_price_day),
                precio_total_dias
            ]
            ws_budget.append(row)
            for cell in ws_budget[ws_budget.max_row]:
                cell.font = cell_font
                cell.alignment = alignment_center
                cell.border = thin_border

        # Ajustar ancho de las columnas
        for column in ws_budget.columns:
            max_length = max(len(str(cell.value)) for cell in column)
            adjusted_width = max_length + 2
            ws_budget.column_dimensions[column[0].column_letter].width = adjusted_width

    # Crear la hoja de resumen general al final
    hoja_resumen_general = plantilla.create_sheet(title="Resumen General")

    # Título del resumen general
    hoja_resumen_general['A1'] = f'PRESUPUESTO: {budget.budget_name} (Resumen General)'
    hoja_resumen_general['A1'].font = Font(bold=True, size=16, name="Calibri")
    hoja_resumen_general['A1'].alignment = alignment_center

    # Encabezados de la tabla de resumen general
    headers = ['ITEM', 'DESCRIPCION DE ITEM', 'UND', 'CANT', 'P.U.', 'SUBTOTAL']
    hoja_resumen_general.append(headers)
    for cell in hoja_resumen_general[2]:  # Segunda fila
        cell.font = header_font
        cell.fill = fill_blue
        cell.alignment = alignment_center
        cell.border = thin_border

    # Datos de los ítems en la hoja de resumen general
    for items in items_by_type.values():
        for item in items:
            precio_total_dias = Decimal(item.precio_item_proyecto)
            row = [
                item.inventario.num_articulo,
                item.inventario.descripcion,
                item.unidad_medida,
                item.cantidad,
                Decimal(item.real_price_day),
                precio_total_dias
            ]
            hoja_resumen_general.append(row)
            for cell in hoja_resumen_general[hoja_resumen_general.max_row]:
                cell.font = cell_font
                cell.alignment = alignment_center
                cell.border = thin_border

    # Añadir el resumen financiero al final de la hoja de resumen general
    hoja_resumen_general.append([''])
    resumen = [
        ('TOTAL PARCIAL', f'S/ {budget.total_soles_parcial:.2f}'),
        ('GASTOS ADMINISTRATIVOS', f'S/ {budget.total_gastos_administrativos:.2f}'),
        ('UTILIDAD', f'S/ {budget.total_utilidad:.2f}'),
        ('TOTAL FINAL', f'S/ {budget.total_soles_final:.2f}'),
        ('TOTAL EN DÓLARES', f'$ {budget.total_dolares_final:.2f}')
    ]
    for label, value in resumen:
        hoja_resumen_general.append([label, '', '', '', '', value])
        for cell in hoja_resumen_general[hoja_resumen_general.max_row]:
            cell.font = cell_font if label != 'TOTAL EN DÓLARES' else Font(bold=True, size=12, color="FF0000")
            cell.alignment = alignment_center
            cell.border = thin_border
            if label == 'TOTAL EN DÓLARES':
                cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    # Crear la respuesta HTTP con el nombre del archivo según el nombre del proyecto
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="cotizacion_{project.project_name}.xlsx"'
    plantilla.save(response)
    return response
